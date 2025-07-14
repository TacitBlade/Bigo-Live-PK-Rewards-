import streamlit as st
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
import tempfile
import os
import pandas as pd
import matplotlib.pyplot as plt

# Read valid PK data from sheet
def read_rules_rewards(sheet):
    pk_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pk_type, pk_points, win_reward = row[0], row[1], row[2]

        if pk_type is None or pk_points is None or win_reward is None:
            continue
        if not isinstance(pk_points, (int, float)):
            continue

        diamonds_needed = int(pk_points / 10)
        pk_data.append({
            "pk_type": pk_type,
            "pk_points": pk_points,
            "win_reward": win_reward,
            "diamonds_needed": diamonds_needed
        })
    return pk_data

# Determine most efficient PK option
def calculate_efficiency(pk_data, diamonds_available):
    options = [pk for pk in pk_data if pk["diamonds_needed"] <= diamonds_available]
    if not options:
        return None
    return max(options, key=lambda x: x["win_reward"] / x["diamonds_needed"])

# Create styled Excel breakdown
def generate_excel(optimal_pk, diamonds_available):
    wb = Workbook()
    ws = wb.active
    ws.title = "PK Efficiency"

    header_font = Font(bold=True)
    ws.append(["Diamonds Available", diamonds_available])
    ws.append(["PK Type", optimal_pk["pk_type"]])
    ws.append(["PK Points", optimal_pk["pk_points"]])
    ws.append(["Diamonds Needed", optimal_pk["diamonds_needed"]])
    ws.append(["Win Reward", optimal_pk["win_reward"]])
    ws.append(["Efficiency (Reward/Diamond)", round(optimal_pk["win_reward"] / optimal_pk["diamonds_needed"], 2)])

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    return temp_file.name

# Generate requirements file
def generate_requirements():
    with open("requirements.txt", "w") as f:
        f.write("\n".join(["streamlit", "openpyxl", "pandas", "matplotlib"]))

# ----- Streamlit UI -----
st.set_page_config(page_title="PK Reward Optimizer", layout="centered")
st.title("💎 PK Reward Efficiency Calculator")

uploaded_file = st.file_uploader("📤 Upload your event Excel file (.xlsx)")
diamonds = st.number_input("Enter the number of diamonds you have", min_value=0, value=0, step=1)

if uploaded_file and diamonds:
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        if "Rules and rewards" in wb.sheetnames:
            sheet = wb["Rules and rewards"]
            pk_data = read_rules_rewards(sheet)
            optimal = calculate_efficiency(pk_data, diamonds)

            df = pd.DataFrame(pk_data)
            df["Efficiency"] = df["win_reward"] / df["diamonds_needed"]
            df_filtered = df[df["diamonds_needed"] <= diamonds]

            # Bar chart
            st.subheader("📊 Reward Efficiency by PK Type")
            fig, ax = plt.subplots()
            ax.bar(df_filtered["pk_type"], df_filtered["Efficiency"], color="mediumseagreen")
            ax.set_ylabel("Reward per Diamond")
            ax.set_xlabel("PK Type")
            ax.set_title("Efficiency Overview")
            st.pyplot(fig)

            # Data table
            st.subheader("📋 Available PK Options")
            st.dataframe(df_filtered[["pk_type", "diamonds_needed", "win_reward", "Efficiency"]]
                         .sort_values(by="Efficiency", ascending=False),
                         use_container_width=True)

            # Optimal result
            if optimal:
                st.success("🏆 Most Efficient PK Reward Found:")
                st.markdown(f"""
                    - **PK Type**: {optimal['pk_type']}  
                    - **Diamonds Needed**: {optimal['diamonds_needed']}  
                    - **Win Reward**: {optimal['win_reward']}  
                    - **Efficiency**: {round(optimal['win_reward'] / optimal['diamonds_needed'], 2)} reward/diamond
                """)

                file_path = generate_excel(optimal, diamonds)
                with open(file_path, "rb") as f:
                    st.download_button("📥 Download Excel Breakdown", f, file_name="pk_efficiency_breakdown.xlsx")

                os.remove(file_path)
            else:
                st.warning("No eligible PK found for your diamond amount.")
        else:
            st.error("Sheet 'Rules and rewards' not found in uploaded file.")
    except Exception as e:
        st.error(f"Error processing file: {e}")

# Generate requirements file
generate_requirements()